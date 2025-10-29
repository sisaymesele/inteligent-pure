from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import plotly.graph_objects as go
from plotly.offline import plot

from management_project.models import InitiativePlanning
from management_project.forms import InitiativePlanningForm
from management_project.services.permission import role_required, get_user_permissions

# -------------------- INITIATIVE LIST --------------------
@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='initiative_planning', action='view')
def initiative_planning_list(request):
    query = request.GET.get('search', '').strip()
    selected_focus = request.GET.get('initiative_focus_area', '').strip()
    page_number = request.GET.get('page', 1)

    initiatives = InitiativePlanning.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by focus area
    if selected_focus:
        initiatives = initiatives.filter(initiative_focus_area=selected_focus)

    # Search filter across multiple fields
    if query:
        initiatives = initiatives.filter(
            Q(initiative_focus_area__icontains=query) |
            Q(initiative_dimension__icontains=query) |
            Q(initiative_name__icontains=query) |
            Q(description__icontains=query)
        )

    initiatives = initiatives.order_by('priority', 'risk_level', '-id')

    # Get distinct focus areas for dropdown
    focus_areas = InitiativePlanning.objects.filter(
        organization_name=request.user.organization_name
    ).values_list('initiative_focus_area', flat=True).distinct()

    paginator = Paginator(initiatives, 10)
    page_obj = paginator.get_page(page_number)

    permissions = get_user_permissions(request.user)

    return render(request, 'initiative_planning/list.html', {
        'initiatives': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'focus_areas': focus_areas,
        'selected_focus': selected_focus,
        'permissions': permissions,
    })


# -------------------- CREATE INITIATIVE --------------------
@login_required
@role_required(['editor', 'owner', 'admin'], model_name='initiative_planning', action='create')
def create_initiative_planning(request):
    """
    Create a new initiative_planning for the current user's organization.
    Redirects back to the child form if 'next' is provided,
    optionally preselecting the newly created initiative_planning.
    """
    next_url = request.GET.get("next") or request.POST.get("next")

    if request.method == "POST":
        form = InitiativePlanningForm(request.POST, request=request)

        # Check if this is a field update (focus area or dimension change)
        if 'focus_area_update' in request.POST or 'dimension_update' in request.POST:
            # This is a field update - just redisplay the form with current data
            if form.is_valid():
                # Form is valid but we don't save, just redisplay
                messages.info(request, "Form updated - click Save to create initiative")
            # Return the form with current data (whether valid or not)
            permissions = get_user_permissions(request.user)
            return render(request, "initiative_planning/form.html", {
                "form": form,
                "next": next_url,
                "permissions": permissions,
            })

        # This is an actual save operation
        if form.is_valid():
            initiative = form.save(commit=False)
            initiative.organization_name = request.user.organization_name
            initiative.save()
            form.save_m2m()  # Save many-to-many relationships (aligned objectives)
            messages.success(request, "Initiative Planning created successfully!")

            # Redirect back to next_url if provided
            if next_url:
                separator = '&' if '?' in next_url else '?'
                return redirect(f"{next_url}{separator}initiative_planning={initiative.pk}")

            return redirect("initiative_planning_list")  # fallback to initiative list

    else:
        form = InitiativePlanningForm(request=request)

    permissions = get_user_permissions(request.user)

    return render(request, "initiative_planning/form.html", {
        "form": form,
        "next": next_url,
        "permissions": permissions,
    })


# -------------------- UPDATE INITIATIVE --------------------
@login_required
@role_required(['editor', 'owner', 'admin'], model_name='initiative_planning', action='edit')
def update_initiative_planning(request, pk):
    initiative = get_object_or_404(
        InitiativePlanning, pk=pk, organization_name=request.user.organization_name
    )

    if request.method == "POST":
        form = InitiativePlanningForm(request.POST, instance=initiative, request=request)

        # Check if this is a field update (focus area or dimension change)
        if 'focus_area_update' in request.POST or 'dimension_update' in request.POST:
            # This is a field update - just redisplay the form with current data
            if form.is_valid():
                # Form is valid but we don't save, just redisplay
                messages.info(request, "Form updated - click Update to save changes")
            # Return the form with current data (whether valid or not)
            permissions = get_user_permissions(request.user)
            context = {
                "form": form,
                "form_title": "Update Initiative Planning",
                "submit_button_text": "Update Initiative Planning",
                "edit_mode": True,
                "editing_initiative": initiative,
                "next": request.GET.get("next", None),
                "permissions": permissions,
            }
            return render(request, 'initiative_planning/form.html', context)

        # This is an actual save operation
        if form.is_valid():
            form.save()
            messages.success(request, "Initiative Planning updated successfully!")
            return redirect("initiative_planning_list")
    else:
        form = InitiativePlanningForm(instance=initiative, request=request)

    permissions = get_user_permissions(request.user)

    context = {
        "form": form,
        "form_title": "Update Initiative Planning",
        "submit_button_text": "Update Initiative Planning",
        "edit_mode": True,
        "editing_initiative": initiative,
        "next": request.GET.get("next", None),
        "permissions": permissions,
    }
    return render(request, "initiative_planning/form.html", context)


# -------------------- DELETE INITIATIVE --------------------
@login_required
@role_required(['owner', 'admin'], model_name='initiative_planning', action='delete')
def delete_initiative_planning(request, pk):
    # Get the initiative belonging to the logged-in user's organization
    initiative = get_object_or_404(
        InitiativePlanning,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        initiative.delete()
        messages.success(request, "Initiative Planning deleted successfully!")
        return redirect('initiative_planning_list')

    permissions = get_user_permissions(request.user)

    # Render confirmation page for GET requests
    return render(request, 'initiative_planning/delete_confirm.html', {
        'entry': initiative,
        'permissions': permissions,
    })


# -------------------- EXPORT INITIATIVE TO EXCEL --------------------
@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='initiative_planning', action='view')
def export_initiative_planning_to_excel(request):
    query = request.GET.get('search', '').strip()
    selected_focus = request.GET.get('initiative_focus_area', '').strip()

    initiatives = InitiativePlanning.objects.filter(
        organization_name=request.user.organization_name
    )

    if selected_focus:
        initiatives = initiatives.filter(initiative_focus_area=selected_focus)
    if query:
        initiatives = initiatives.filter(
            Q(initiative_focus_area__icontains=query) |
            Q(initiative_dimension__icontains=query) |
            Q(initiative_name__icontains=query) |
            Q(description__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initiative Planning"

    # === Styles ===
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # === Headers ===
    headers = [
        "Focus Area", "Dimension", "Name", "Description",
        "Priority", "Impact", "Likelihood", "Risk Level",
        "Baseline Status", "Target Status", "Total Budget Planned",
        "Total HR Planned", "Aligned Objectives", "Start Date", "End Date",
        "Created At", "Updated At"
    ]

    total_columns = len(headers)

    # === Title Row ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1, value="Initiative Planning")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill

    # === Header Row ===
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # === Data Rows ===
    for row_num, ini in enumerate(initiatives, start=3):
        aligned_objs = ", ".join([obj.objective for obj in ini.aligned_objectives.all()])
        row_data = [
            ini.initiative_focus_area,
            ini.initiative_dimension,
            ini.initiative_name,
            ini.description or "",
            ini.get_priority_display(),
            ini.get_impact_display(),
            ini.get_likelihood_display(),
            ini.get_risk_level_display(),
            ini.get_baseline_status_display(),
            ini.get_target_status_display(),
            float(ini.total_budget_planned),
            float(ini.total_hr_planned),
            aligned_objs,
            ini.start_date.strftime("%Y-%m-%d") if ini.start_date else "",
            ini.end_date.strftime("%Y-%m-%d") if ini.end_date else "",
            ini.created_at.strftime("%Y-%m-%d %H:%M"),
            ini.updated_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            # Optional: Format numbers
            if col_num in [11, 12]:  # Budget & HR
                cell.number_format = '#,##0.00'

    # === Auto column width ===
    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 5, 50)

    # === Response ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Initiative_Planning.xlsx'
    wb.save(response)
    return response


# -------------------- INITIATIVE CHART --------------------
@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='initiative_planning', action='view')
def initiative_planning_chart(request):
    qs = InitiativePlanning.objects.filter(
        organization_name=request.user.organization_name
    )

    LEVEL_ORDER = ['Very High', 'High', 'Medium', 'Low', 'Very Low']
    LEVEL_COLORS = {
        'Very High': '#d62728',  # red
        'High': '#ff7f0e',       # orange
        'Medium': '#ffbb78',     # light orange
        'Low': '#98df8a',        # light green
        'Very Low': '#2ca02c',   # green
    }

    # Common axis configuration
    axis_config = dict(
        showline=True,
        linewidth=2,
        linecolor='black',
        mirror=True,
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        zeroline=True,
        zerolinewidth=2,
        zerolinecolor='gray'
    )

    # ------------------ Helper: Bar Chart ------------------
    def create_bar_chart(field_name, title):
        counts = qs.values(field_name).annotate(count=Count('id'))
        data_dict = {lvl: 0 for lvl in LEVEL_ORDER}
        for c in counts:
            key = c[field_name] if c[field_name] else 'Medium'
            if key in data_dict:
                data_dict[key] = c['count']

        fig = go.Figure(data=[go.Bar(
            x=list(data_dict.keys()),
            y=list(data_dict.values()),
            text=list(data_dict.values()),
            textposition='auto',
            marker_color=[LEVEL_COLORS[k] for k in data_dict.keys()],
            hovertemplate='<b>%{x}</b><br>Count: %{y}<extra></extra>'
        )])
        fig.update_layout(
            title=dict(text=title, x=0.5, xanchor='center', font=dict(size=16)),
            xaxis=dict(
                title=field_name.replace("_", " ").title(),
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            yaxis=dict(
                title='Count',
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            height=400,
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=60, r=40, t=80, b=80)
        )
        return plot(fig, output_type='div', include_plotlyjs=False)

    # ------------------ Charts ------------------
    plot_priority = create_bar_chart('priority', 'Distribution by Priority')
    plot_impact = create_bar_chart('impact', 'Distribution by Impact')
    plot_likelihood = create_bar_chart('likelihood', 'Distribution by Likelihood')
    plot_risk = create_bar_chart('risk_level', 'Distribution by Risk Level')

    # Status Pie Chart
    status_counts = qs.values('baseline_status').annotate(count=Count('id'))
    status_summary = [{'status': s['baseline_status'], 'count': s['count']} for s in status_counts]

    fig_status = go.Figure(data=[go.Pie(
        labels=[s['status'] for s in status_summary],
        values=[s['count'] for s in status_summary],
        textinfo='label+percent',
        hole=0.3,
        hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>'
    )])
    fig_status.update_layout(
        title=dict(text='Distribution by Status', x=0.5, xanchor='center', font=dict(size=16)),
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        margin=dict(l=60, r=40, t=80, b=80)
    )
    plot_status = plot(fig_status, output_type='div', include_plotlyjs=False)

    # Dimension Chart
    dimension_counts = qs.values('initiative_dimension').annotate(count=Count('id')).order_by('-count')
    fig_dimension = go.Figure(data=[go.Bar(
        x=[d['initiative_dimension'] for d in dimension_counts],
        y=[d['count'] for d in dimension_counts],
        text=[d['count'] for d in dimension_counts],
        textposition='auto',
        marker_color=['#636efa', '#ef553b', '#00cc96', '#ab63fa', '#ffa15a'],
        hovertemplate='<b>%{x}</b><br>Count: %{y}<extra></extra>'
    )])
    fig_dimension.update_layout(
        title=dict(text='Count per Dimension', x=0.5, xanchor='center', font=dict(size=16)),
        xaxis=dict(
            title='Dimension',
            **axis_config,
            title_font=dict(size=14, color='black'),
            tickfont=dict(size=12, color='black')
        ),
        yaxis=dict(
            title='Count',
            **axis_config,
            title_font=dict(size=14, color='black'),
            tickfont=dict(size=12, color='black')
        ),
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        margin=dict(l=60, r=40, t=80, b=80)
    )
    plot_dimension = plot(fig_dimension, output_type='div', include_plotlyjs=False)

    # Focus Area Chart
    focus_counts = qs.values('initiative_focus_area').annotate(count=Count('id')).order_by('-count')
    fig_focus = go.Figure(data=[go.Bar(
        y=[d['initiative_focus_area'] for d in focus_counts],
        x=[d['count'] for d in focus_counts],
        text=[d['count'] for d in focus_counts],
        textposition='auto',
        orientation='h',
        marker_color='#4F81BD',
        hovertemplate='<b>%{y}</b><br>Count: %{x}<extra></extra>'
    )])
    fig_focus.update_layout(
        title=dict(text='Count per Focus Area', x=0.5, xanchor='center', font=dict(size=16)),
        yaxis=dict(
            title='Focus Area',
            **axis_config,
            title_font=dict(size=14, color='black'),
            tickfont=dict(size=12, color='black')
        ),
        xaxis=dict(
            title='Count',
            **axis_config,
            title_font=dict(size=14, color='black'),
            tickfont=dict(size=12, color='black')
        ),
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        margin=dict(l=80, r=40, t=80, b=80)
    )
    plot_focus = plot(fig_focus, output_type='div', include_plotlyjs=False)

    # ------------------ Summary Cards ------------------
    total_initiatives = qs.count()
    priority_summary = [{'level': lvl, 'count': qs.filter(priority=lvl).count(), 'color': LEVEL_COLORS[lvl]} for lvl in LEVEL_ORDER]

    # Additional stats
    total_budget = sum(float(ini.total_budget_planned) for ini in qs)
    total_hr = sum(float(ini.total_hr_planned) for ini in qs)
    high_risk_count = qs.filter(risk_level__in=['Very High', 'High']).count()

    stats = {
        'total_initiatives': total_initiatives,
        'total_budget': total_budget,
        'total_hr': total_hr,
        'high_risk_count': high_risk_count,
        'high_risk_percent': round((high_risk_count / total_initiatives * 100), 1) if total_initiatives > 0 else 0,
    }

    permissions = get_user_permissions(request.user)

    return render(request, 'initiative_planning/chart.html', {
        'plot_priority': plot_priority,
        'plot_impact': plot_impact,
        'plot_likelihood': plot_likelihood,
        'plot_risk': plot_risk,
        'plot_status': plot_status,
        'plot_dimension': plot_dimension,
        'plot_focus': plot_focus,
        'total_initiatives': total_initiatives,
        'priority_summary': priority_summary,
        'stats': stats,
        'permissions': permissions,
    })