from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import calendar
from django.core.paginator import Paginator
from django.urls import reverse

from management_project.models import SwotReport, StrategicCycle
from management_project.forms import SwotReportForm
from management_project.services.permission import role_required, get_user_permissions

from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.db.models import Count, Q
import plotly.graph_objects as go


# -------------------- SWOT LIST --------------------
@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_report', action='view')
def swot_report_list(request):
    query = request.GET.get("search", "").strip()
    cycle_slug = request.GET.get("cycle", "").strip()
    page_number = request.GET.get("page", 1)

    org = request.user.organization_name

    # Get all SWOT reports for this organization
    swots = SwotReport.objects.filter(organization_name=org)

    # Filter by strategic cycle if provided
    strategic_cycle = None
    if cycle_slug:
        strategic_cycle = StrategicCycle.objects.filter(
            slug=cycle_slug, organization_name=org
        ).first()
        if strategic_cycle:
            swots = swots.filter(strategic_report_period=strategic_cycle)  # Fixed: strategic_report_period

    # Apply search filter
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query) |
            Q(strategic_report_period__name__icontains=query) |  # This is correct
            Q(strategic_report_period__time_horizon__icontains=query) |  # This is correct
            Q(strategic_report_period__time_horizon_type__icontains=query)  # This is correct
        )

    # Ordering
    swots = swots.order_by("strategic_report_period__end_date", "swot_type", "priority", "-created_at")

    # Pagination
    paginator = Paginator(swots, 10)  # 10 items per page
    page_obj = paginator.get_page(page_number)

    # Available strategic cycles for filter dropdown
    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=org
    ).order_by('-start_date')

    permissions = get_user_permissions(request.user)

    return render(request, "swot_report/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "strategic_cycles": strategic_cycles,
        "selected_cycle": cycle_slug,
        "permissions": permissions,
    })

# -------------------- CREATE SWOT --------------------


# -------------------- CREATE SWOT --------------------
@login_required
@role_required(['editor', 'owner', 'admin'], model_name='swot_report', action='create')
def create_swot_report(request):
    org = request.user.organization_name
    # Get the latest StrategicCycle for this organization
    latest_cycle = StrategicCycle.objects.filter(organization_name=org).order_by('-start_date').first()

    if request.method == 'POST':
        form = SwotReportForm(request.POST, request=request)
        if 'save' in request.POST and form.is_valid():
            swot_entry = form.save(commit=False)
            swot_entry.organization_name = org
            # Assign the selected strategic report or fallback to latest
            if not swot_entry.strategic_report_period:  # Fixed: use strategic_report_period
                swot_entry.strategic_report_period = latest_cycle
            swot_entry.save()
            messages.success(request, "SWOT entry created successfully!")
            return redirect('swot_report_list')

    else:
        form = SwotReportForm(request=request)
        if latest_cycle:
            form.fields['strategic_report_period'].initial = latest_cycle

    permissions = get_user_permissions(request.user)
    return render(request, 'swot_report/form.html', {
        'form': form,
        'permissions': permissions,
    })


# -------------------- UPDATE SWOT --------------------
@login_required
@role_required(['editor', 'owner', 'admin'], model_name='swot_report', action='edit')
def update_swot_report(request, pk):
    entry = get_object_or_404(
        SwotReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = SwotReportForm(request.POST, instance=entry, request=request)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "SWOT entry updated successfully!")
            return redirect('swot_report_list')

    else:
        form = SwotReportForm(instance=entry, request=request)

    permissions = get_user_permissions(request.user)
    return render(request, 'swot_report/form.html', {
        'form': form,
        'permissions': permissions,
    })


# -------------------- DELETE SWOT --------------------
@login_required
@role_required(['owner', 'admin'], model_name='swot_report', action='delete')
def delete_swot_report(request, pk):
    entry = get_object_or_404(
        SwotReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "SWOT entry deleted successfully!")
        return redirect('swot_report_list')

    permissions = get_user_permissions(request.user)
    return render(request, 'swot_report/delete_confirm.html', {
        'entry': entry,
        'permissions': permissions,
    })
#



@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_report', action='view')
def swot_report_export_to_excel(request):
    org_profile = request.user.organization_name
    query = request.GET.get("search", "").strip()
    cycle_slug = request.GET.get("cycle", "").strip()
    page_number = request.GET.get("page", 1)
    export = request.GET.get("export", "").lower()

    # -------------------- BASE QUERY --------------------
    swots = SwotReport.objects.filter(
        organization_name=org_profile
    ).select_related('strategic_report_period')

    # -------------------- STRATEGIC CYCLE FILTER --------------------
    strategic_cycle_name = ""
    has_specific_cycle = False
    if cycle_slug:
        try:
            cycle = StrategicCycle.objects.get(slug=cycle_slug, organization_name=org_profile)
            strategic_cycle_name = cycle.name
            has_specific_cycle = True
            swots = swots.filter(strategic_report_period=cycle)
        except StrategicCycle.DoesNotExist:
            pass

    # -------------------- SEARCH FILTER --------------------
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query) |
            Q(strategic_report_period__name__icontains=query) |
            Q(strategic_report_period__time_horizon__icontains=query) |
            Q(strategic_report_period__time_horizon_type__icontains=query)
        )

    # -------------------- EXCEL EXPORT --------------------
    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "SWOT Reports"

        # Styles
        title_font = Font(size=14, bold=True, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # -------------------- TITLE ROW --------------------
        total_columns = 8 if has_specific_cycle else 9
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"SWOT Report for {strategic_cycle_name}" if strategic_cycle_name else "SWOT Report Export"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = title_fill
        ws.row_dimensions[1].height = 28

        # -------------------- HEADER ROW --------------------
        headers = ["SWOT Type", "Pillar", "Factor", "Description", "Priority", "Impact", "Likelihood", "Created At"]
        if not has_specific_cycle:
            headers = ["Strategic Cycle"] + headers

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[2].height = 22

        # -------------------- DATA ROWS --------------------
        for row_num, s in enumerate(swots, start=3):
            row_data = [
                s.swot_type,
                s.swot_pillar,
                s.swot_factor,
                s.description or "",
                s.priority,
                s.impact,
                s.likelihood or "",
                s.created_at.strftime("%Y-%m-%d"),
            ]
            if not has_specific_cycle:
                row_data = [s.strategic_report_period.name] + row_data

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border

        # -------------------- MINIMIZED COLUMN WIDTH --------------------
        max_widths = [15, 20, 25, 40, 12, 12, 18, 12]  # default widths
        if not has_specific_cycle:
            max_widths = [20] + max_widths  # for strategic cycle column

        for i, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, max_widths[i-1])

        # -------------------- HTTP RESPONSE --------------------
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"SWOT_Report_{strategic_cycle_name.replace(' ', '_')}.xlsx" if strategic_cycle_name else "SWOT_Report_Export.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    # -------------------- PAGINATION & TEMPLATE --------------------
    paginator = Paginator(swots, 10)
    page_obj = paginator.get_page(page_number)

    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=org_profile
    ).order_by('-start_date')
    permissions = get_user_permissions(request.user)

    return render(request, "swot_report/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "strategic_cycles": strategic_cycles,
        "selected_cycle": cycle_slug,
        "permissions": permissions,
    })


# -------------------- SWOT CHARTS --------------------
# @login_required
# @role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_report', action='view')
# def swot_report_chart(request):
#     qs = SwotReport.objects.filter(
#         organization_name=request.user.organization_name
#     )
#
#     # Optional: filter by strategic cycle from GET parameter
#     cycle_filter = request.GET.get("cycle")
#     if cycle_filter:
#         qs = qs.filter(strategic_report_period__name=cycle_filter)
#
#     # ------------------ COMMON LAYOUT SETTINGS ------------------
#     layout_style = dict(
#         template='plotly_white',
#         height=420,
#         margin=dict(l=60, r=40, t=80, b=60),
#         title_font=dict(size=18, color="#333333", family="Arial Black"),
#         xaxis=dict(
#             title_font=dict(size=14, color="#333333"),
#             showgrid=True,
#             gridcolor='rgba(200,200,200,0.3)',
#             zeroline=True,
#             zerolinecolor='rgba(0,0,0,0.4)',
#             tickangle=0,
#             showline=True,
#             linecolor='rgba(0,0,0,0.6)',
#         ),
#         yaxis=dict(
#             title_font=dict(size=14, color="#333333"),
#             showgrid=True,
#             gridcolor='rgba(200,200,200,0.3)',
#             zeroline=True,
#             zerolinecolor='rgba(0,0,0,0.4)',
#             showline=True,
#             linecolor='rgba(0,0,0,0.6)',
#         ),
#     )
#
#     # ------------------ SWOT TYPE DISTRIBUTION ------------------
#     type_counts = qs.values('swot_type').annotate(count=Count('id'))
#     type_labels = [t['swot_type'] for t in type_counts]
#     type_values = [t['count'] for t in type_counts]
#
#     fig_swot_type = go.Figure(data=[go.Bar(
#         x=type_labels,
#         y=type_values,
#         text=type_values,
#         textposition='outside',
#         marker_color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728'],
#         hovertemplate='%{x}<br>Count: %{y}<extra></extra>'
#     )])
#     fig_swot_type.update_layout(
#         title=f"SWOT Distribution by Type ({cycle_filter or 'All Cycles'})",
#         xaxis_title="SWOT Type",
#         yaxis_title="Count",
#         **layout_style
#     )
#
#     # ------------------ PRIORITY DISTRIBUTION ------------------
#     priority_counts = qs.values('priority').annotate(count=Count('id'))
#     priority_labels = [p['priority'] for p in priority_counts]
#     priority_values = [p['count'] for p in priority_counts]
#
#     fig_priority = go.Figure(data=[go.Bar(
#         x=priority_labels,
#         y=priority_values,
#         text=priority_values,
#         textposition='outside',
#         marker_color=['#d62728', '#ff7f0e', '#ffbb78', '#98df8a', '#2ca02c'],
#         hovertemplate='%{x}<br>Count: %{y}<extra></extra>'
#     )])
#     fig_priority.update_layout(
#         title=f"SWOT Distribution by Priority ({cycle_filter or 'All Cycles'})",
#         xaxis_title="Priority",
#         yaxis_title="Count",
#         **layout_style
#     )
#
#     # ------------------ IMPACT DISTRIBUTION ------------------
#     impact_counts = qs.values('impact').annotate(count=Count('id'))
#     impact_labels = [i['impact'] for i in impact_counts]
#     impact_values = [i['count'] for i in impact_counts]
#
#     fig_impact = go.Figure(data=[go.Bar(
#         x=impact_labels,
#         y=impact_values,
#         text=impact_values,
#         textposition='outside',
#         marker_color=['#98df8a', '#2ca02c', '#ffbb78', '#ff7f0e', '#d62728'],
#         hovertemplate='%{x}<br>Count: %{y}<extra></extra>'
#     )])
#     fig_impact.update_layout(
#         title=f"SWOT Distribution by Impact ({cycle_filter or 'All Cycles'})",
#         xaxis_title="Impact",
#         yaxis_title="Count",
#         **layout_style
#     )
#
#     # ------------------ LIKELIHOOD DISTRIBUTION ------------------
#     likelihood_counts = qs.values('likelihood').annotate(count=Count('id'))
#     likelihood_labels = [l['likelihood'] if l['likelihood'] else 'Unknown' for l in likelihood_counts]
#     likelihood_values = [l['count'] for l in likelihood_counts]
#
#     fig_likelihood = go.Figure(data=[go.Bar(
#         x=likelihood_labels,
#         y=likelihood_values,
#         text=likelihood_values,
#         textposition='outside',
#         marker_color=['#17becf', '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#7f7f7f'],
#         hovertemplate='%{x}<br>Count: %{y}<extra></extra>'
#     )])
#     fig_likelihood.update_layout(
#         title=f"SWOT Distribution by Likelihood ({cycle_filter or 'All Cycles'})",
#         xaxis_title="Likelihood",
#         yaxis_title="Count",
#         **layout_style
#     )
#
#     # ------------------ PILLAR ANALYSIS ------------------
#     pillar_counts = qs.values('swot_pillar').annotate(count=Count('id')).order_by('-count')
#     pillar_labels = [p['swot_pillar'] for p in pillar_counts]
#     pillar_values = [p['count'] for p in pillar_counts]
#
#     fig_pillar = go.Figure(data=[go.Bar(
#         x=pillar_labels,
#         y=pillar_values,
#         text=pillar_values,
#         textposition='outside',
#         marker_color=['#636efa', '#ef553b', '#00cc96', '#ab63fa', '#ffa15a'],
#         hovertemplate='%{x}<br>Count: %{y}<extra></extra>'
#     )])
#     fig_pillar.update_layout(
#         title=f"SWOT Count per Pillar ({cycle_filter or 'All Cycles'})",
#         xaxis_title="Pillar",
#         yaxis_title="Count",
#         **layout_style
#     )
#
#     # ------------------ SUMMARY DATA ------------------
#     total_swot = qs.count()
#     summary_data = {
#         'total_swot': total_swot,
#         'high_priority_count': qs.filter(priority__in=['High', 'Very High']).count(),
#         'strengths_count': qs.filter(swot_type='Strength').count(),
#         'weaknesses_count': qs.filter(swot_type='Weakness').count(),
#         'opportunities_count': qs.filter(swot_type='Opportunity').count(),
#         'threats_count': qs.filter(swot_type='Threat').count(),
#         'cycle': cycle_filter or 'All Cycles'
#     }
#
#     # Dropdown Cycles
#     all_cycles = qs.values_list('strategic_report_period__name', flat=True).distinct()
#
#     permissions = get_user_permissions(request.user)
#
#     return render(request, 'swot_report/chart.html', {
#         'plot_html_swot_type': fig_swot_type.to_html(full_html=False),
#         'plot_html_priority': fig_priority.to_html(full_html=False),
#         'plot_html_impact': fig_impact.to_html(full_html=False),
#         'plot_html_likelihood': fig_likelihood.to_html(full_html=False),
#         'plot_html_pillar': fig_pillar.to_html(full_html=False),
#         'summary_data': summary_data,
#         'all_cycles': all_cycles,
#         'selected_cycle': cycle_filter,
#         'permissions': permissions,
#     })


# -------------------- SWOT CHARTS --------------------
@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_report', action='view')
def swot_report_chart(request):
    org = request.user.organization_name
    qs = SwotReport.objects.filter(organization_name=org)

    # Get all available strategic cycles for dropdown
    all_cycles = StrategicCycle.objects.filter(
        organization_name=org
    ).order_by('-start_date')

    # Filter by strategic cycle from GET parameter
    cycle_filter = request.GET.get("cycle")
    selected_cycle_name = "All Cycles"
    selected_cycle_obj = None

    if cycle_filter:
        try:
            selected_cycle_obj = StrategicCycle.objects.get(
                slug=cycle_filter,
                organization_name=org
            )
            qs = qs.filter(strategic_report_period=selected_cycle_obj)
            selected_cycle_name = selected_cycle_obj.name
        except StrategicCycle.DoesNotExist:
            pass

    # Handle empty data case
    if qs.count() == 0:
        empty_chart_html = """
        <div class="text-center py-5">
            <i class="bi bi-bar-chart text-muted" style="font-size: 3rem;"></i>
            <h5 class="text-muted mt-3">No SWOT Data Available</h5>
            <p class="text-muted">No SWOT reports found for the selected criteria.</p>
        </div>
        """
        return render(request, 'swot_report/chart.html', {
            'plot_html_swot_type': empty_chart_html,
            'plot_html_priority': empty_chart_html,
            'plot_html_impact': empty_chart_html,
            'plot_html_likelihood': empty_chart_html,
            'plot_html_pillar': empty_chart_html,
            'summary_data': {
                'total_swot': 0,
                'high_priority_count': 0,
                'strengths_count': 0,
                'weaknesses_count': 0,
                'opportunities_count': 0,
                'threats_count': 0,
                'cycle': selected_cycle_name
            },
            'all_cycles': all_cycles,
            'selected_cycle': cycle_filter,
            'selected_cycle_obj': selected_cycle_obj,
            'permissions': get_user_permissions(request.user),
        })

    # ------------------ COMMON LAYOUT SETTINGS ------------------
    layout_style = dict(
        template='plotly_white',
        height=450,
        margin=dict(l=60, r=40, t=100, b=80),
        title_font=dict(size=20, color="#2c3e50", family="Arial Black"),
        font=dict(family="Arial", size=12),
        plot_bgcolor='rgba(248,249,250,1)',
        paper_bgcolor='rgba(255,255,255,1)',
        xaxis=dict(
            title_font=dict(size=14, color="#2c3e50", family="Arial"),
            showgrid=True,
            gridcolor='rgba(200,200,200,0.3)',
            zeroline=True,
            zerolinecolor='rgba(0,0,0,0.2)',
            tickangle=0,
            showline=True,
            linecolor='rgba(0,0,0,0.3)',
        ),
        yaxis=dict(
            title_font=dict(size=14, color="#2c3e50", family="Arial"),
            showgrid=True,
            gridcolor='rgba(200,200,200,0.3)',
            zeroline=True,
            zerolinecolor='rgba(0,0,0,0.2)',
            showline=True,
            linecolor='rgba(0,0,0,0.3)',
        ),
        hoverlabel=dict(
            bgcolor="white",
            font_size=12,
            font_family="Arial"
        )
    )

    # ------------------ SWOT TYPE DISTRIBUTION ------------------
    type_counts = qs.values('swot_type').annotate(count=Count('id')).order_by('swot_type')
    type_labels = [t['swot_type'] for t in type_counts]
    type_values = [t['count'] for t in type_counts]

    # Color mapping for SWOT types
    swot_colors = {
        'Strength': '#2ecc71',  # Green
        'Weakness': '#e74c3c',  # Red
        'Opportunity': '#3498db',  # Blue
        'Threat': '#f39c12'  # Orange
    }

    type_colors = [swot_colors.get(t, '#95a5a6') for t in type_labels]

    fig_swot_type = go.Figure(data=[go.Bar(
        x=type_labels,
        y=type_values,
        text=type_values,
        textposition='auto',
        textfont=dict(size=14, color='white', weight='bold'),
        marker_color=type_colors,
        hovertemplate='<b>%{x}</b><br>Count: %{y}<extra></extra>',
        marker_line=dict(width=1, color='rgba(0,0,0,0.2)'),
    )])
    fig_swot_type.update_layout(
        title=dict(
            text=f"SWOT Type Distribution<br><sub>{selected_cycle_name}</sub>",
            x=0.5,
            xanchor='center'
        ),
        xaxis_title="SWOT Type",
        yaxis_title="Number of Items",
        **layout_style
    )

    # ------------------ PRIORITY DISTRIBUTION ------------------
    priority_order = ['Very Low', 'Low', 'Medium', 'High', 'Very High']
    priority_counts = qs.values('priority').annotate(count=Count('id'))
    priority_dict = {p['priority']: p['count'] for p in priority_counts}
    priority_labels = [p for p in priority_order if p in priority_dict]
    priority_values = [priority_dict[p] for p in priority_labels]

    priority_colors = ['#27ae60', '#2ecc71', '#f39c12', '#e67e22', '#e74c3c']  # Green to Red

    fig_priority = go.Figure(data=[go.Bar(
        x=priority_labels,
        y=priority_values,
        text=priority_values,
        textposition='auto',
        textfont=dict(size=14, color='white', weight='bold'),
        marker_color=priority_colors,
        hovertemplate='<b>%{x} Priority</b><br>Count: %{y}<extra></extra>',
        marker_line=dict(width=1, color='rgba(0,0,0,0.2)'),
    )])
    fig_priority.update_layout(
        title=dict(
            text=f"Priority Distribution<br><sub>{selected_cycle_name}</sub>",
            x=0.5,
            xanchor='center'
        ),
        xaxis_title="Priority Level",
        yaxis_title="Number of Items",
        **layout_style
    )

    # ------------------ IMPACT DISTRIBUTION ------------------
    impact_order = ['Very Low', 'Low', 'Medium', 'High', 'Very High']
    impact_counts = qs.values('impact').annotate(count=Count('id'))
    impact_dict = {i['impact']: i['count'] for i in impact_counts}
    impact_labels = [i for i in impact_order if i in impact_dict]
    impact_values = [impact_dict[i] for i in impact_labels]

    impact_colors = ['#27ae60', '#2ecc71', '#f39c12', '#e67e22', '#e74c3c']

    fig_impact = go.Figure(data=[go.Bar(
        x=impact_labels,
        y=impact_values,
        text=impact_values,
        textposition='auto',
        textfont=dict(size=14, color='white', weight='bold'),
        marker_color=impact_colors,
        hovertemplate='<b>%{x} Impact</b><br>Count: %{y}<extra></extra>',
        marker_line=dict(width=1, color='rgba(0,0,0,0.2)'),
    )])
    fig_impact.update_layout(
        title=dict(
            text=f"Impact Distribution<br><sub>{selected_cycle_name}</sub>",
            x=0.5,
            xanchor='center'
        ),
        xaxis_title="Impact Level",
        yaxis_title="Number of Items",
        **layout_style
    )

    # ------------------ LIKELIHOOD DISTRIBUTION ------------------
    likelihood_order = ['Very Low', 'Low', 'Medium', 'High', 'Very High']
    likelihood_counts = qs.values('likelihood').annotate(count=Count('id'))
    likelihood_dict = {l['likelihood']: l['count'] for l in likelihood_counts if l['likelihood']}
    likelihood_labels = [l for l in likelihood_order if l in likelihood_dict]
    likelihood_values = [likelihood_dict[l] for l in likelihood_labels]

    if not likelihood_labels:  # Handle case where all likelihoods are null
        likelihood_labels = ['Unknown']
        likelihood_values = [qs.filter(likelihood__isnull=True).count()]

    likelihood_colors = ['#27ae60', '#2ecc71', '#f39c12', '#e67e22', '#e74c3c'][:len(likelihood_labels)]

    fig_likelihood = go.Figure(data=[go.Bar(
        x=likelihood_labels,
        y=likelihood_values,
        text=likelihood_values,
        textposition='auto',
        textfont=dict(size=14, color='white', weight='bold'),
        marker_color=likelihood_colors,
        hovertemplate='<b>%{x} Likelihood</b><br>Count: %{y}<extra></extra>',
        marker_line=dict(width=1, color='rgba(0,0,0,0.2)'),
    )])
    fig_likelihood.update_layout(
        title=dict(
            text=f"Likelihood Distribution<br><sub>{selected_cycle_name}</sub>",
            x=0.5,
            xanchor='center'
        ),
        xaxis_title="Likelihood Level",
        yaxis_title="Number of Items",
        **layout_style
    )

    # ------------------ PILLAR ANALYSIS ------------------
    pillar_counts = qs.values('swot_pillar').annotate(count=Count('id')).order_by('-count')
    pillar_labels = [p['swot_pillar'] for p in pillar_counts]
    pillar_values = [p['count'] for p in pillar_counts]

    fig_pillar = go.Figure(data=[go.Bar(
        x=pillar_labels,
        y=pillar_values,
        text=pillar_values,
        textposition='auto',
        textfont=dict(size=14, color='white', weight='bold'),
        marker_color=['#3498db', '#9b59b6', '#1abc9c', '#34495e', '#e67e22', '#7f8c8d'],
        hovertemplate='<b>%{x}</b><br>Count: %{y}<extra></extra>',
        marker_line=dict(width=1, color='rgba(0,0,0,0.2)'),
    )])
    fig_pillar.update_layout(
        title=dict(
            text=f"Pillar Analysis<br><sub>{selected_cycle_name}</sub>",
            x=0.5,
            xanchor='center'
        ),
        xaxis_title="Strategic Pillar",
        yaxis_title="Number of Items",
        **layout_style
    )

    # ------------------ SUMMARY DATA ------------------
    total_swot = qs.count()
    summary_data = {
        'total_swot': total_swot,
        'high_priority_count': qs.filter(priority__in=['High', 'Very High']).count(),
        'strengths_count': qs.filter(swot_type='Strength').count(),
        'weaknesses_count': qs.filter(swot_type='Weakness').count(),
        'opportunities_count': qs.filter(swot_type='Opportunity').count(),
        'threats_count': qs.filter(swot_type='Threat').count(),
        'cycle': selected_cycle_name,
        'cycle_obj': selected_cycle_obj
    }

    permissions = get_user_permissions(request.user)

    return render(request, 'swot_report/chart.html', {
        'plot_html_swot_type': fig_swot_type.to_html(full_html=False, include_plotlyjs=False),
        'plot_html_priority': fig_priority.to_html(full_html=False, include_plotlyjs=False),
        'plot_html_impact': fig_impact.to_html(full_html=False, include_plotlyjs=False),
        'plot_html_likelihood': fig_likelihood.to_html(full_html=False, include_plotlyjs=False),
        'plot_html_pillar': fig_pillar.to_html(full_html=False, include_plotlyjs=False),
        'summary_data': summary_data,
        'all_cycles': all_cycles,
        'selected_cycle': cycle_filter,
        'selected_cycle_obj': selected_cycle_obj,
        'permissions': permissions,
    })