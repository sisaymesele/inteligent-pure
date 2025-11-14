from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from management_project.models import SwotAnalysis
from management_project.forms import SwotAnalysisForm
from management_project.services.permission import role_required, get_user_permissions
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Count
import plotly.graph_objects as go
from django.utils import timezone


@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_analysis', action='view')
def swot_analysis_list(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('swot_type', '').strip()
    page_number = request.GET.get('page', 1)

    # Base queryset: all SWOT analyses for this organization
    swots = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by SWOT type if selected
    if selected_type:
        swots = swots.filter(swot_type=selected_type)

    # Search filter across multiple fields
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query)
        )

    # Convert queryset to list for custom sorting
    swots_list = list(swots)

    # Define custom orderings
    swot_type_order = {'Strength': 1, 'Weakness': 2, 'Opportunity': 3, 'Threat': 4}
    priority_order = {'High': 1, 'Medium': 2, 'Low': 3}
    impact_order = {'High': 1, 'Medium': 2, 'Low': 3}
    likelihood_order = {'Almost Certain': 1, 'Likely': 2, 'Possible': 3, 'Unlikely': 4}

    # Sort list
    swots_list.sort(key=lambda x: (
        swot_type_order.get(x.swot_type, 5),
        priority_order.get(x.priority, 99),
        impact_order.get(x.impact, 99),
        likelihood_order.get(x.likelihood, 99),
        -x.created_at.timestamp()  # newest first
    ))

    # Pagination
    paginator = Paginator(swots_list, 10)
    page_obj = paginator.get_page(page_number)

    # Provide SWOT types for dropdown filter
    swot_types = [choice[0] for choice in SwotAnalysis.SWOT_TYPES]

    permissions = get_user_permissions(request.user)

    return render(request, 'swot_analysis/list.html', {
        'swots': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'swot_types': swot_types,
        'selected_type': selected_type,
        'permissions': permissions,
    })



# -------------------- CREATE SWOT --------------------
@login_required
@role_required(['editor', 'owner', 'admin'], model_name='swot_analysis', action='create')
def create_swot_analysis(request):
    if request.method == 'POST':
        form = SwotAnalysisForm(request.POST, request=request)  # Pass request here
        # Only save if the Save button is clicked
        if 'save' in request.POST and form.is_valid():
            swot_entry = form.save(commit=False)
            swot_entry.organization_name = request.user.organization_name
            swot_entry.save()
            messages.success(request, "SWOT entry created successfully!")
            return redirect('swot_analysis_list')
    else:
        form = SwotAnalysisForm(request=request)  # Pass request here

    permissions = get_user_permissions(request.user)

    return render(request, 'swot_analysis/form.html', {
        'form': form,
        'permissions': permissions,
    })


# -------------------- UPDATE SWOT --------------------
@login_required
@role_required(['editor', 'owner', 'admin'], model_name='swot_analysis', action='edit')
def update_swot_analysis(request, pk):
    # Fetch the entry only if it belongs to the user's organization
    entry = get_object_or_404(SwotAnalysis, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = SwotAnalysisForm(request.POST, instance=entry, request=request)  # Pass request here
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "SWOT entry updated successfully!")
            return redirect('swot_analysis_list')
        else:
            messages.error(request, "Error updating SWOT entry. Please check the form.")
    else:
        form = SwotAnalysisForm(instance=entry, request=request)  # Pass request here

    permissions = get_user_permissions(request.user)

    return render(request, 'swot_analysis/form.html', {
        'form': form,
        'permissions': permissions,
    })


# -------------------- DELETE SWOT --------------------
@login_required
@role_required(['owner', 'admin'], model_name='swot_analysis', action='delete')
def delete_swot_analysis(request, pk):
    entry = get_object_or_404(
        SwotAnalysis,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "SWOT entry deleted successfully!")
        return redirect('swot_analysis_list')

    permissions = get_user_permissions(request.user)

    return render(request, 'swot_analysis/delete_confirm.html', {
        'entry': entry,
        'permissions': permissions,
    })



# -------------------- EXPORT SWOT TO EXCEL --------------------


@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_analysis', action='view')
def export_swot_analysis_to_excel(request):
    # Get filters from request
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('swot_type', '').strip()

    # Base queryset
    swots = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # Apply filters
    if selected_type:
        swots = swots.filter(swot_type=selected_type)
    if query:
        search_filter = (
                Q(swot_type__icontains=query) |
                Q(swot_pillar__icontains=query) |
                Q(swot_factor__icontains=query) |
                Q(description__icontains=query)
        )
        swots = swots.filter(search_filter)

    # Convert queryset to list for custom sorting (SAME ORDERING AS LIST VIEW)
    swots_list = list(swots)

    # Define custom orderings (SAME AS LIST VIEW)
    swot_type_order = {'Strength': 1, 'Weakness': 2, 'Opportunity': 3, 'Threat': 4}
    priority_order = {'Very High': 1, 'High': 2, 'Medium': 3, 'Low': 4, 'Very Low': 5}
    impact_order = {'Very High': 1, 'High': 2, 'Medium': 3, 'Low': 4, 'Very Low': 5}
    likelihood_order = {'Almost Certain': 1, 'Likely': 2, 'Possible': 3, 'Unlikely': 4, 'Rare': 5}

    # Sort list (SAME LOGIC AS LIST VIEW)
    swots_list.sort(key=lambda x: (
        swot_type_order.get(x.swot_type, 5),
        priority_order.get(x.priority, 99),
        impact_order.get(x.impact, 99),
        likelihood_order.get(x.likelihood, 99),
        -x.created_at.timestamp()  # newest first
    ))

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SWOT Analysis"

    # MATCHED COLOR SCHEME (Same as Web Template)
    COLORS = {
        # SWOT Types (Matches your badge colors)
        'strength_bg': '28A745',  # bg-success - Green
        'weakness_bg': 'DC3545',  # bg-danger - Red
        'opportunity_bg': 'FFC107',  # bg-warning - Amber/Yellow
        'threat_bg': '343A40',  # bg-dark - Dark Gray

        # Priority Levels (Matches your badge colors)
        'very_high_priority': 'DC3545',  # bg-danger - Red
        'high_priority': 'FFC107',  # bg-warning - Amber/Yellow
        'medium_priority': '007BFF',  # bg-primary - Blue
        'low_priority': '28A745',  # bg-success - Green
        'very_low_priority': '17A2B8',  # bg-info - Light Blue

        # Impact Levels (Same as Priority for consistency)
        'very_high_impact': 'DC3545',  # bg-danger - Red
        'high_impact': 'FFC107',  # bg-warning - Amber/Yellow
        'medium_impact': '007BFF',  # bg-primary - Blue
        'low_impact': '28A745',  # bg-success - Green
        'very_low_impact': '17A2B8',  # bg-info - Light Blue

        # Likelihood Levels
        'almost_certain': 'DC3545',  # bg-danger - Red
        'likely': 'FFC107',  # bg-warning - Amber/Yellow
        'possible': '007BFF',  # bg-primary - Blue
        'unlikely': '28A745',  # bg-success - Green
        'rare': '17A2B8',  # bg-info - Light Blue

        # Layout Colors
        'title_bg': '2C3E50',
        'header_bg': '34495E',
        'alt_row1': 'F8F9FA',
        'alt_row2': 'E9ECEF',
        'border': 'BDC3C7',
    }

    # Font Styles with optimized text colors
    title_font = Font(size=16, bold=True, color="FFFFFF", name='Calibri')
    header_font = Font(size=14, bold=True, color="FFFFFF", name='Calibri')
    data_font = Font(size=13, color="2C3E50", name='Calibri')
    white_font = Font(size=13, bold=True, color="FFFFFF", name='Calibri')
    dark_font = Font(size=13, bold=True, color="000000", name='Calibri')  # For light backgrounds
    meta_font = Font(size=10, italic=True, color="7F8C8D", name='Calibri')  # Increased size for metadata

    # Fills
    title_fill = PatternFill(start_color=COLORS['title_bg'], end_color=COLORS['title_bg'], fill_type="solid")
    header_fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
    alt_fill1 = PatternFill(start_color=COLORS['alt_row1'], end_color=COLORS['alt_row1'], fill_type="solid")
    alt_fill2 = PatternFill(start_color=COLORS['alt_row2'], end_color=COLORS['alt_row2'], fill_type="solid")
    meta_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")

    # Border
    elegant_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    # Title row
    total_columns = 9
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "SWOT ANALYSIS RESULT"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = title_fill
    ws.row_dimensions[1].height = 35

    # FIXED: Metadata row with better text handling
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    meta_cell = ws.cell(row=2, column=1)

    # Create metadata text
    record_count = len(swots_list)
    generated_date = timezone.now().strftime('%B %d, %Y at %H:%M:%S')
    meta_text = f"Generated on {generated_date} | Total Records: {record_count}"

    meta_cell.value = meta_text
    meta_cell.font = meta_font
    meta_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Added wrap_text
    meta_cell.fill = meta_fill
    meta_cell.border = elegant_border

    # FIXED: Increased metadata row height to prevent text hiding
    ws.row_dimensions[2].height = 30  # Increased from 22 to 30

    # Header row
    headers = [
        "SWOT Type", "Strategic Pillar", "Key Factor", "Detailed Description",
        "Priority Level", "Impact Score", "Likelihood",
        "Created Date", "Last Updated"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = elegant_border
    ws.row_dimensions[3].height = 35

    # Data rows with optimized color coding and text handling
    # USING THE SORTED swots_list INSTEAD OF ORIGINAL swots QUERYSET
    for row_num, swot in enumerate(swots_list, start=4):
        # Format dates
        created_date = swot.created_at.strftime("%B, %Y") if swot.created_at else ""
        updated_date = swot.updated_at.strftime("%B, %Y") if swot.updated_at else ""

        row_data = [
            swot.swot_type,
            swot.swot_pillar,
            swot.swot_factor,
            swot.description or "",
            swot.priority,
            swot.impact,
            swot.likelihood or "",
            created_date,
            updated_date,
        ]

        # Alternate row colors
        row_fill = alt_fill1 if row_num % 2 == 0 else alt_fill2

        # FIXED: Improved dynamic row height calculation
        base_height = 25
        description = str(swot.description or "")
        pillar = str(swot.swot_pillar or "")
        factor = str(swot.swot_factor or "")

        # Calculate lines needed for each field
        max_lines = 1

        # Description field (column 4) - more generous line calculation
        if description:
            # Use 55 characters per line for description (reduced from 60)
            desc_lines = (len(description) // 55) + 1
            max_lines = max(max_lines, min(desc_lines, 10))  # Cap at 10 lines

        # Pillar field (column 2)
        if pillar:
            pillar_lines = (len(pillar) // 25) + 1  # Reduced from 30
            max_lines = max(max_lines, min(pillar_lines, 5))

        # Factor field (column 3)
        if factor:
            factor_lines = (len(factor) // 25) + 1  # Reduced from 30
            max_lines = max(max_lines, min(factor_lines, 5))

        # Calculate dynamic height with better minimum handling
        dynamic_height = 25 + (max(1, max_lines) - 1) * 14  # Reduced line spacing from 15 to 14
        base_height = min(max(dynamic_height, 25), 100)  # Ensure minimum 25, maximum 100

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.fill = row_fill
            cell.border = elegant_border

            # FIXED: Improved alignment with better text handling
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="left",
                shrink_to_fit=False
            )

            # OPTIMIZED COLOR CODING (Matches Web Template)

            # Column 1: SWOT Type
            if col_num == 1:
                swot_type_colors = {
                    'Strength': COLORS['strength_bg'],
                    'Weakness': COLORS['weakness_bg'],
                    'Opportunity': COLORS['opportunity_bg'],
                    'Threat': COLORS['threat_bg'],
                }
                color = swot_type_colors.get(str(value).strip(), None)
                if color:
                    # Use white text for dark backgrounds, black text for light backgrounds
                    if value == 'Opportunity':  # Yellow background
                        cell.font = dark_font
                    else:  # Green, Red, Dark backgrounds
                        cell.font = white_font
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Column 5: Priority
            elif col_num == 5:
                priority_colors = {
                    'Very High': COLORS['very_high_priority'],
                    'High': COLORS['high_priority'],
                    'Medium': COLORS['medium_priority'],
                    'Low': COLORS['low_priority'],
                    'Very Low': COLORS['very_low_priority'],
                }
                color = priority_colors.get(str(value).strip(), None)
                if color:
                    # Use white text for dark backgrounds, black text for light backgrounds
                    if value in ['High']:  # Yellow background
                        cell.font = dark_font
                    else:  # Red, Blue, Green, Light Blue backgrounds
                        cell.font = white_font
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Column 6: Impact
            elif col_num == 6:
                impact_colors = {
                    'Very High': COLORS['very_high_impact'],
                    'High': COLORS['high_impact'],
                    'Medium': COLORS['medium_impact'],
                    'Low': COLORS['low_impact'],
                    'Very Low': COLORS['very_low_impact'],
                }
                color = impact_colors.get(str(value).strip(), None)
                if color:
                    if value in ['High']:  # Yellow background
                        cell.font = dark_font
                    else:
                        cell.font = white_font
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Column 7: Likelihood
            elif col_num == 7:
                likelihood_colors = {
                    'Almost Certain': COLORS['almost_certain'],
                    'Likely': COLORS['likely'],
                    'Possible': COLORS['possible'],
                    'Unlikely': COLORS['unlikely'],
                    'Rare': COLORS['rare'],
                }
                color = likelihood_colors.get(str(value).strip(), None)
                if color:
                    if value in ['Likely']:  # Yellow background
                        cell.font = dark_font
                    else:
                        cell.font = white_font
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Columns 8 & 9: Dates - center align
            elif col_num in [8, 9]:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        # Set row height
        ws.row_dimensions[row_num].height = base_height

    # FIXED: Adjusted column widths for better text visibility
    column_widths = {
        1: 16,  # SWOT Type (increased from 14)
        2: 28,  # Strategic Pillar (increased from 25)
        3: 28,  # Key Factor (increased from 25)
        4: 55,  # Detailed Description (increased from 50)
        5: 16,  # Priority Level (increased from 14)
        6: 16,  # Impact Score (increased from 14)
        7: 18,  # Likelihood (increased from 16)
        8: 18,  # Created Date (increased from 16)
        9: 18,  # Last Updated (increased from 16)
    }

    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)

        for cell in col:
            if cell.value:
                text = str(cell.value)
                # More conservative line length calculation
                if len(text) > 45:
                    avg_line_length = 45
                    max_length = max(max_length, avg_line_length)
                else:
                    max_length = max(max_length, len(text))

        preferred_width = column_widths.get(i, min(max_length + 4, 65))  # Increased max width
        ws.column_dimensions[column_letter].width = preferred_width

    # Freeze panes and add auto-filter
    ws.freeze_panes = "A4"
    if swots_list:  # Use swots_list instead of swots
        ws.auto_filter.ref = f"A3:{get_column_letter(total_columns)}{3 + len(swots_list)}"

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="SWOT_Analysis_Result_{timezone.now().strftime("%B_%Y")}.xlsx"'
    wb.save(response)
    return response


# -------------------- SWOT CHARTS --------------------
@login_required
@role_required(['viewer', 'editor', 'owner', 'admin'], model_name='swot_analysis', action='view')
def swot_analysis_chart(request):
    qs = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # ------------------ SWOT TYPE DISTRIBUTION ------------------
    type_counts = qs.values('swot_type').annotate(count=Count('id'))
    type_labels = [t['swot_type'] for t in type_counts]
    type_values = [t['count'] for t in type_counts]

    fig_swot_type = go.Figure(data=[go.Bar(
        x=type_labels,
        y=type_values,
        text=type_values,
        textposition='auto',
        marker_color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']  # consistent colors
    )])
    fig_swot_type.update_layout(
        title='SWOT Distribution by Type',
        xaxis_title='SWOT Type',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ PRIORITY DISTRIBUTION ------------------
    priority_counts = qs.values('priority').annotate(count=Count('id'))
    priority_labels = [p['priority'] for p in priority_counts]
    priority_values = [p['count'] for p in priority_counts]

    fig_priority = go.Figure(data=[go.Bar(
        x=priority_labels,
        y=priority_values,
        text=priority_values,
        textposition='auto',
        marker_color=['#d62728', '#ff7f0e', '#ffbb78', '#98df8a', '#2ca02c']
    )])
    fig_priority.update_layout(
        title='SWOT Distribution by Priority',
        xaxis_title='Priority',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ IMPACT DISTRIBUTION ------------------
    impact_counts = qs.values('impact').annotate(count=Count('id'))
    impact_labels = [i['impact'] for i in impact_counts]
    impact_values = [i['count'] for i in impact_counts]

    fig_impact = go.Figure(data=[go.Bar(
        x=impact_labels,
        y=impact_values,
        text=impact_values,
        textposition='auto',
        marker_color=['#98df8a', '#2ca02c', '#ffbb78', '#ff7f0e', '#d62728']
    )])
    fig_impact.update_layout(
        title='SWOT Distribution by Impact',
        xaxis_title='Impact',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ LIKELIHOOD DISTRIBUTION ------------------
    likelihood_counts = qs.values('likelihood').annotate(count=Count('id'))
    likelihood_labels = [l['likelihood'] if l['likelihood'] else 'Unknown' for l in likelihood_counts]
    likelihood_values = [l['count'] for l in likelihood_counts]

    fig_likelihood = go.Figure(data=[go.Bar(
        x=likelihood_labels,
        y=likelihood_values,
        text=likelihood_values,
        textposition='auto',
        marker_color=['#17becf', '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#7f7f7f']
    )])
    fig_likelihood.update_layout(
        title='SWOT Distribution by Likelihood',
        xaxis_title='Likelihood',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ PILLAR ANALYSIS ------------------
    pillar_counts = qs.values('swot_pillar').annotate(count=Count('id')).order_by('-count')
    pillar_labels = [p['swot_pillar'] for p in pillar_counts]
    pillar_values = [p['count'] for p in pillar_counts]

    fig_pillar = go.Figure(data=[go.Bar(
        x=pillar_labels,
        y=pillar_values,
        text=pillar_values,
        textposition='auto',
        marker_color=['#636efa', '#ef553b', '#00cc96', '#ab63fa', '#ffa15a']
    )])
    fig_pillar.update_layout(
        title='SWOT Count per Pillar',
        xaxis_title='Pillar',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ SUMMARY DATA ------------------
    total_swot = qs.count()
    high_priority_count = qs.filter(priority__in=['High', 'Very High']).count()
    strengths_count = qs.filter(swot_type='Strength').count()
    weaknesses_count = qs.filter(swot_type='Weakness').count()
    opportunities_count = qs.filter(swot_type='Opportunity').count()
    threats_count = qs.filter(swot_type='Threat').count()

    summary_data = {
        'total_swot': total_swot,
        'high_priority_count': high_priority_count,
        'strengths_count': strengths_count,
        'weaknesses_count': weaknesses_count,
        'opportunities_count': opportunities_count,
        'threats_count': threats_count,
    }

    permissions = get_user_permissions(request.user)

    return render(request, 'swot_analysis/chart.html', {
        'plot_html_swot_type': fig_swot_type.to_html(full_html=False),
        'plot_html_priority': fig_priority.to_html(full_html=False),
        'plot_html_impact': fig_impact.to_html(full_html=False),
        'plot_html_likelihood': fig_likelihood.to_html(full_html=False),
        'plot_html_pillar': fig_pillar.to_html(full_html=False),
        'summary_data': summary_data,
        'permissions': permissions,
    })
